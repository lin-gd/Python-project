import openpyxl
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, TimeoutException
import time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys


# Open the spreadsheet and get the latest dues status.
wb = openpyxl.load_workbook('DueRecords.xlsx')
sheet = wb['DueRecords']

lastCol = sheet.max_column
latestMonth = sheet.cell(row=1, column=lastCol).value
print('The last column in the excel sheet is '+ str(lastCol))

# Check each member's payment status.
unpaidMembers = {}
for r in range(2, sheet.max_row + 1):
    payment = sheet.cell(row=r, column=lastCol).value
    if payment != 'paid':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaidMembers[name] = email
print('The members who haven’t paid in the most recent month :')
print(unpaidMembers)

#Logs into the email account
driver = webdriver.Firefox(executable_path="C:\\Users\\username\\.\\.\\.\\geckodriver.exe")
driver.get('http://www.gmail.com')

driver.find_element_by_name('identifier').send_keys('youremail')
driver.find_element_by_class_name('VfPpkd-RLmnJb').click()

time.sleep(3)

pswd=driver.find_element_by_name("password")
pswd.send_keys('yourpassword')
driver.find_element_by_class_name('VfPpkd-RLmnJb').click()

time.sleep(3)
# Send out reminder emails.
def send_email(email_to, email_subject, email_name):
    try:
        composeElem=driver.find_element_by_xpath('/html/body/div[7]/div[3]/div/div[2]/div[1]/div[1]/div[1]/div/div/div/div[1]/div/div')
        composeElem.click()
    except NoSuchElementException: 
        print("NoSuchElementException occurs")
    
    time.sleep(7)
    try:
        wait=WebDriverWait(driver, 30)
        input_box = wait.until(EC.element_to_be_clickable((By.XPATH, "//textarea[@spellcheck='false' and @autocomplete='false'][@aria-label='To']")))
        input_box.click()
    except TimeoutException:
        # WebDriverWait throws TimeoutException if it fails
        print("TimeoutException occurs")

    #time.sleep(10)
    input_box.send_keys(email_to)
    sub=driver.find_element_by_name('subjectbox')
    sub.send_keys(email_subject)
    sub.send_keys(Keys.TAB)
    driver.switch_to.active_element.send_keys( 'Dear '+ email_name +',Records show that you have not paid dues . Please make this payment as soon as possible. Thank you!')
    driver.switch_to.active_element.click()
    time.sleep(5)
    driver.switch_to.active_element.send_keys(Keys.TAB)
    driver.switch_to.active_element.click()
    # display.stop()

for name, email in unpaidMembers.items():
    send_email(email,'Dues Unpaid',name)
    print('Email Sent! to '+name)
    
driver.close()
#Converting excel to csv
import os,openpyxl,csv
os.chdir("C:\\Users\\username\\.\\..\\..")
wb=openpyxl.load_workbook('DueRecords.xlsx')
sheet=wb['DueRecords']
oFile=open('Dues.csv','w',newline='')
oWriter = csv.writer(oFile)
oWriter.writerows(sheet.values)
print('Excel is converted to Csv')
oFile.close()
with open('Dues.csv',newline='') as csvfile:
    rows = csv.reader(csvfile,delimiter=',')
    for row in rows:
        print(row)
